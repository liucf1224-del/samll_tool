from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (QApplication, QWidget, QPushButton, QLineEdit, QLabel, QFileDialog,
                             QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout)
from openpyxl import load_workbook
import sys

class ExcelReplaceTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 文字替换工具")
        # self.setWindowIcon(QIcon("icon.ico"))  # 设置窗口图标
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # 输入文件
        input_layout = QHBoxLayout()
        self.input_path = QLineEdit()
        btn_input = QPushButton("选择文件")
        btn_input.clicked.connect(self.select_input)
        input_layout.addWidget(QLabel("输入文件:"))
        input_layout.addWidget(self.input_path)
        input_layout.addWidget(btn_input)

        # 输出文件
        output_layout = QHBoxLayout()
        self.output_path = QLineEdit()
        output_layout.addWidget(QLabel("保存的文件名:"))
        output_layout.addWidget(self.output_path)

        # 替换表格
        self.table = QTableWidget(5, 2)
        self.table.setHorizontalHeaderLabels(["原文字", "新文字"])

        # 替换按钮
        replace_btn = QPushButton("开始替换")
        replace_btn.clicked.connect(self.run_replace)

        # 状态栏
        self.status = QLabel("就绪")

        layout.addLayout(input_layout)
        layout.addLayout(output_layout)
        layout.addWidget(self.table)
        layout.addWidget(replace_btn)
        layout.addWidget(self.status)
        self.setLayout(layout)

    def select_input(self):
        fname, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xlsx)")
        if fname:
            self.input_path.setText(fname)

    def get_replacements(self):
        replacements = {}
        for row in range(self.table.rowCount()):
            old_item = self.table.item(row, 0)
            new_item = self.table.item(row, 1)
            if old_item and new_item:
                old_text = old_item.text().strip()
                new_text = new_item.text().strip()
                if old_text and new_text:
                    replacements[old_text] = new_text
        return replacements

    def run_replace(self):
        input_file = self.input_path.text()
        output_file = self.output_path.text()
        replacements = self.get_replacements()

        if not input_file or not output_file or not replacements:
            self.status.setText("请填写完整信息")
            return

            # 自动添加 .xlsx 扩展名（如果没有）
        if not output_file.lower().endswith('.xlsx'):
            output_file += '.xlsx'
            self.output_path.setText(output_file)  # 更新输出文件路径显示

        try:
            self.replace_chinese_in_excel(input_file, output_file, replacements)
            self.status.setText(f"替换完成！保存至 {output_file}")
        except Exception as e:
            self.status.setText(f"发生错误：{str(e)}")

    def replace_chinese_in_excel(self, input_file, output_file, replacements):
        wb = load_workbook(input_file)
        replaced = False  # 标记是否发生了替换
        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        for old_text, new_text in replacements.items():
                            if old_text in cell.value:
                                cell.value = cell.value.replace(old_text, new_text)
                                replaced = True  # 只要有一个替换，标记为 Tr
                            # 如果替换了内容，更新单元格值（可选）
                            if cell.value != original_value:
                                cell.value = cell.value
        if not replaced:
            raise Exception("未找到任何匹配的原文字，替换未执行")
        wb.save(output_file)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    tool = ExcelReplaceTool()
    tool.show()
    sys.exit(app.exec_())
